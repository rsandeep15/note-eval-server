import warnings,requests, xml.etree.ElementTree as ET, openpyxl
from openpyxl.cell import column_index_from_string
import urllib, requests
import json, datetime
from time import mktime
from string import Template
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class RealEstate:
    # The mapping maps "Street Address", "Zip", "Zillow" and "Trulia" to column numbers
    def writeToFirebaseDatabase(self, sheet, user, start, mapping):
        #Authentication steps
        firebase_database_keyfile = open('/var/www/API_Keys' +
        '/firebase_database_key.txt', "r")
        firebase_database_key = firebase_database_keyfile.read().strip('\n')

        endpoint_template= Template('https://loannotesassistant.firebaseio.com'+
        '/users/$user/notes.json?auth=$authKey')
        endpoint = endpoint_template.substitute(user=user,
        authKey=firebase_database_key)

        doc = {}
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                column = column_index_from_string(cell.column)
                key_cell = sheet.cell(row = start, column = column).value
                if isinstance(key_cell, basestring):
                    key_cell = key_cell.replace("\n", "").strip()
                #Get a set of values row-column wise
                value_cell = cell.value
                if isinstance(value_cell, basestring):
                    value_cell = value_cell.replace("\n", "")
                    # Do not enter the heading into the database
                    if key_cell == value_cell:
                        continue
                if isinstance(value_cell, datetime.datetime):
                    value_cell = int(mktime(value_cell.timetuple()))
                #Zillow API call to get current price
                if key_cell == "Zillow":
                    addressColumn = column_index_from_string(mapping["Street Address"])
                    zipcodeColumn = column_index_from_string(mapping["Zip"])
                    address = sheet.cell(row = cell.row, column = addressColumn).value
                    zipcode = sheet.cell(row = cell.row, column = zipcodeColumn).value
                    zillowPrice = self.getZillowPrice(address, zipcode)
                    if zillowPrice != "Not found":
                        value_cell = zillowPrice
                doc[key_cell] = value_cell
            if doc["Street Address"] is not None:
                addEntry = requests.post(endpoint, data = json.dumps(doc))
    def validateMapping(self, mapping, row):
        keys = mapping.keys()
        if "Street Address" in keys and "Zip" in keys and "Zillow" in keys and "Trulia" in keys:
            mapping["Start"] = row
            return True
        return False
    # Finds all the columns with the vital information needed to query APIs
    def findColumns(self, sheet):
        mapping = {}
        keys = ["Street Address", "Zip", "Zillow", "Trulia", "Start"]
        for row in sheet.iter_rows():
            # Trimming possible white spaces on edge before comparison
            for cell in row:
                if cell.value is not None and str(cell.value).strip() in keys:
                    mapping[str(cell.value).strip()] = cell.column
            if self.validateMapping(mapping, cell.row):
                return mapping
        # Returns a dictionary of the mappings for look up
        return "Mapping Failed"

    # Queries Zillow API for a price of a home given its address and zipcode
    def getZillowPrice(self, address, zipcode):
        zillow_key_file = open('/var/www/API_Keys' +
        '/zillow_key.txt', "r")
        ZILLOW_API_KEY = zillow_key_file.read().strip('\n')
        ZILLOW_URL = 'http://www.zillow.com/webservice/GetSearchResults.htm'

        data = {'zws-id':ZILLOW_API_KEY, 'address':address,
            'citystatezip':zipcode}

        r = requests.post(ZILLOW_URL, params=data)

        root = ET.fromstring(r.text)
        element = root.find('./response/results/result/zestimate/amount')
        if element is None:
            return "Not found"
        else:
            if str(element.text) == "None":
                return "Not found"
            else:
                return "$" + str(element.text)

    # Queries Trulia API for a price of a home given its address and zipcode
    def getTruliaPrice(self, driver, address, zipcode):
        TRULIA_URL = "http://trulia.com"
        wait = WebDriverWait(driver, 20)
        driver.get(TRULIA_URL)
        element = wait.until(EC.presence_of_element_located ((By.ID,"searchbox_form_location")))
        element.clear()
        element.send_keys(str(address) + " " + str(zipcode))
        element.send_keys(Keys.RETURN)
        price = "Not Found"
        try:
            content = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.typeDeemphasize > span:nth-child(1) > span:nth-child(2)")))
            price = "$" + content.text
        finally:
            return price

    def run(self, publicURL, user):
        # Pull the remote file into a local directory
        FILE_PATH = "data/data.xlsx"
        urllib.urlretrieve(publicURL, FILE_PATH)

        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook('data/data.xlsx', data_only=True)
        sheets = wb.get_sheet_names()
        for sheet_name in sheets:
            sheet = wb.get_sheet_by_name(sheet_name)
            columnsMap = self.findColumns(sheet);
            if (columnsMap == "Mapping Failed"):
                print "Excel sheet: " + sheet_name +  " could not be parsed."
            else:
                address_col = columnsMap["Street Address"]
                zipcode_col = columnsMap["Zip"]

                zillow_col = columnsMap["Zillow"]
                trulia_col = columnsMap["Trulia"]

                start = columnsMap["Start"]
                self.writeToFirebaseDatabase(sheet, user, start, columnsMap)
        #
        # driver = webdriver.Firefox()
        # for x in range(4, sheet.max_row):
        #     address = sheet[address_col + str(x)].value
        #     zipcode = sheet[zipcode_col + str(x)].value
        #
        # #Skip invalid addresses and zipcodes
        #     if address is None or zipcode is None:
        #         continue
        # #
        #     zillowPrice = getZillowPrice(address, zipcode)
        # #     truliaPrice = getTruliaPrice(driver, address, zipcode)
        # #
        #     if zillowPrice != "Not found":
        #         sheet[zillow_col + str(x)] = zillowPrice
        # if truliaPrice != "Not found":
        #     sheet[trulia_col + str(x)] = truliaPrice
        #
        #     output = str(address).ljust(30) + str(zipcode).ljust(15)
        # # output += str(zillowPrice).ljust(30) + str(truliaPrice)
        #     print output
        #
        # wb.save('data/data.xlsx')
